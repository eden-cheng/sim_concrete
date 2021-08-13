import openpyxl
import relation

class Reprocess():
    def __init__(self, feature, ex_case_dic, wb, ws, excel_path, yaml_path, ws_name):
        self.ex_case_dic = ex_case_dic
        self.r_obj = relation.Relation(feature, excel_path, yaml_path, ws_name)
        self.r_obj.func()
        self.ws = ws
        self.wb = wb
        self.excel_path = excel_path

    def func(self):
        self.replace_value()
        self.replace_summary()

    def replace_value(self):
        ######################## 后处理之替换tv_init_speed ########################
        for case_id, value_group in self.ex_case_dic.items():
            print(case_id)
            if not value_group: continue
            for value_sub in value_group:
                for key, value in value_sub.items():
                    temp = str(value).split(' ')
                    temp_key = key.split('_')
                    if temp[0] == 'tv_init_speed':
                        if temp_key[1] == 'tv1':
                            # print(value)
                            # print(value_sub)
                            # print(value_sub['para_tv1_init_speed'])
                            value_sub[key] = value_sub['para_tv1_init_speed']
                            # print(value_sub[key])
                        if temp_key[0] == 'tv2':
                            # print(value_sub['para_tv2_init_speed'])
                            value_sub[key] = value_sub['para_tv2_init_speed']
                # print(value_sub)

    def replace_summary(self):
        ######################## 后处理之替换summary ########################
        for case_id, value_group in self.ex_case_dic.items():
            if not value_group: break
            # print("@@@@@@@: %s" % case_id)
            # print(value_group[0])
            # self.func_parser(case_id, value_group)    #替换summary和rm中的公式部分
            self.form_relpace(value_group)
            # print(value_group[0])
            # print(value_group)
            repalce_para = self.func_replace(case_id)
            # print(repalce_para)
            if not repalce_para: continue
            for value_sub in value_group:
                for target, source in repalce_para.items():
                    # print(target)
                    # print(source)
                    # value = value_sub[source]
                    value = source
                    # print(value)
                    value_sub['summary'] = value_sub['summary'].replace(target, value)
                # print(value_sub['summary'])

    def func_replace(self, case_id):
        cell = self.ws.cell(row=self.r_obj.case_row_dic[case_id], column=self.r_obj.repalce_col_num).value
        if not cell: return
        line = cell.split("\n")
        dic = {}
        for i in line:
            temp = i.split(":")
            dic[temp[0]] = temp[1]
        return dic  

    def func_parser(self, case_id, value_group):
        """
        替换summary和rm中的公式部分
        这里其实不是根据公式的结构去替换的，而是直接将case_lib中的内容，顺序替换到para_lib中
        """
        wb = openpyxl.load_workbook(self.excel_path)
        ws_case = wb['case_lib']
        relation_case = {}
        for cell in ws_case['A'][:200]:
            if cell.value == case_id:
                # print(cell.value)
                # print(case_id)
                summary_content = ws_case.cell(row=cell.row, column=4).value
                rm_content = ws_case.cell(row=cell.row, column=18).value
                # print(summary_content)
                if value_group:
                    for case_dic in value_group:
                        case_dic['summary'] = summary_content
                        case_dic['rm'] = rm_content

    def form_parser(self, formula_value):
        case_lib = self.wb['case_lib']
        temp = formula_value.split('!')
        # print(temp[1])
        data = case_lib[temp[1]].value
        return data

    def form_relpace(self, dic_arr):
        # print(dic_arr[0]['summary'])
        summary_data = self.form_parser(dic_arr[0]['summary'])
        rm_data = self.form_parser(dic_arr[0]['rm'])
        for dic in dic_arr:
            dic['summary'] = summary_data
            dic['rm'] = rm_data